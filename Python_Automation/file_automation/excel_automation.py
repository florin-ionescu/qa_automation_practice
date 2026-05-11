# resource https://www.youtube.com/watch?v=7YS6YDQKFh0
# pip install openpyxl in terminal

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('lista_date.xlsx') #wb = workbook
ws = wb.active #ws = worksheet
print(ws['A1'].value) # you can access the value with .value
ws['A2'].value = 'Jan' #to assign a new value
wb.save('lista_date.xlsx')  #save the file // if excel is open you get error

for row in range(1, 11):
    for column in range(0, 4):
        char = get_column_letter(col) #get_column_letter builtin func

ws.merge_cells('A1:D1') #merge cells
ws.unmerge_cells('A1:D1') #unmerge data, you will lose the data.
ws.insert_cols(2) #func that inserts a col
ws.insert_rows(2) #func that inserts a row

# how to create a new excel file
wb.create_sheet('NewSheet')

create_wb = Workbook()
create_ws = create_wb.active
create_ws.title = "Data"  #name of sheet
create_ws.append(["Use", "Append", "In excel", "Sheet"])

create_wb.save('test_sheet.xlsx') #name of excel file





